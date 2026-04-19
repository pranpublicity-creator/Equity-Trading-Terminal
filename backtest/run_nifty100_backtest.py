"""
NIFTY 100 Pattern Backtest Runner
==================================
Runs all 17 chart-pattern detectors across 96 NIFTY-100 symbols
using locally-cached 15-min OHLCV data (no live Fyers connection needed).

Outputs:
  reports/NIFTY100_Pattern_Backtest_<date>.xlsx   (9 sheets)

Sheets:
  1. Summary          overall portfolio metrics
  2. Per Symbol       win-rate / P&L / Sharpe per stock
  3. Per Pattern      win-rate / P&L / profit-factor per pattern
  4. Regime Edge      performance sliced by regime
  5. Pattern×Regime   heat-map: win-rate for each (pattern, regime) combo
  6. Top 20 Trades    best individual trades
  7. Worst 20 Trades  biggest individual losses
  8. All Trades       full trade log
  9. About            methodology notes

Usage:
  cd "C:\\Users\\pc\\.claude\\TRADING DATA\\EQUITY TERMINAL"
  python backtest/run_nifty100_backtest.py
"""
from __future__ import annotations

import os
import sys
import sqlite3
import time
import traceback
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from typing import List, Dict, Optional

import numpy as np
import pandas as pd

# ── path setup ──────────────────────────────────────────────────────────────
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

import config
from core.stock_universe import NIFTY_100
from patterns.pattern_detector import PatternDetector
from signals.regime_detector import detect_regime
from trading.charge_calculator import ChargeCalculator


# ── constants ────────────────────────────────────────────────────────────────
LOOKBACK   = 200        # bars fed to pattern detector
STEP       = 5          # advance window by STEP bars (speed vs granularity)
MIN_BARS   = LOOKBACK + 50
RR_MIN     = 1.0        # minimum acceptable R:R (same as config.MIN_RISK_REWARD)
MAX_DRIFT  = 0.008      # 0.8% entry-drift cap  (config.MAX_ENTRY_DRIFT_PCT)
CAPITAL    = config.TOTAL_TRADING_CAPITAL
SZ_PCT     = 0.10       # size per trade = 10% of capital


# ── data loading ─────────────────────────────────────────────────────────────
def load_symbol(symbol: str, db_path: str) -> Optional[pd.DataFrame]:
    """Load 15-min OHLCV from SQLite cache."""
    try:
        conn = sqlite3.connect(db_path)
        df = pd.read_sql_query(
            "SELECT timestamp, open, high, low, close, volume "
            "FROM ohlcv_cache "
            "WHERE symbol=? AND resolution='15' "
            "ORDER BY timestamp ASC",
            conn, params=(symbol,)
        )
        conn.close()
        if len(df) < MIN_BARS:
            return None
        df["timestamp"] = df["timestamp"].astype(int)
        df = df.reset_index(drop=True)
        return df
    except Exception as e:
        print(f"  [LOAD ERROR] {symbol}: {e}")
        return None


# ── single-trade record ───────────────────────────────────────────────────────
@dataclass
class Trade:
    symbol:       str
    direction:    str
    pattern:      str
    regime:       str
    confidence:   float
    entry_bar:    int
    entry_price:  float
    stop_loss:    float
    target:       float
    exit_bar:     int   = 0
    exit_price:   float = 0.0
    exit_reason:  str   = ""
    qty:          int   = 1
    gross_pnl:    float = 0.0
    charges:      float = 0.0
    net_pnl:      float = 0.0
    hold_bars:    int   = 0
    entry_ts:     int   = 0
    exit_ts:      int   = 0


# ── core backtest per symbol ─────────────────────────────────────────────────
def backtest_symbol(symbol: str, df: pd.DataFrame,
                    detector: PatternDetector,
                    charge_calc: ChargeCalculator) -> List[Trade]:
    trades: List[Trade] = []
    position: Optional[dict] = None
    close = df["close"].values
    high  = df["high"].values
    low   = df["low"].values
    ts    = df["timestamp"].values

    for i in range(LOOKBACK, len(df), STEP if position is None else 1):
        # ── exit logic (check bar by bar when in a trade) ────────────────
        if position is not None:
            for j in range(position["last_bar"] + 1, min(i + 1, len(df))):
                h, l, c = high[j], low[j], close[j]
                reason = ""
                ep = position["exit_price"]
                if position["direction"] == "BUY":
                    if l <= position["sl"]:   reason, ep = "STOP_LOSS",  position["sl"]
                    elif h >= position["tgt"]: reason, ep = "TARGET",     position["tgt"]
                else:
                    if h >= position["sl"]:   reason, ep = "STOP_LOSS",  position["sl"]
                    elif l <= position["tgt"]: reason, ep = "TARGET",     position["tgt"]

                if not reason and j == len(df) - 1:
                    reason, ep = "END_OF_DATA", c

                if reason:
                    entry = position["entry"]
                    qty   = position["qty"]
                    pnl   = (ep - entry) * qty if position["direction"] == "BUY" \
                            else (entry - ep) * qty
                    chg   = charge_calc.calculate_total(entry, ep, qty)
                    t = Trade(
                        symbol      = symbol,
                        direction   = position["direction"],
                        pattern     = position["pattern"],
                        regime      = position["regime"],
                        confidence  = position["conf"],
                        entry_bar   = position["entry_bar"],
                        entry_price = entry,
                        stop_loss   = position["sl"],
                        target      = position["tgt"],
                        exit_bar    = j,
                        exit_price  = ep,
                        exit_reason = reason,
                        qty         = qty,
                        gross_pnl   = round(pnl, 2),
                        charges     = round(chg, 2),
                        net_pnl     = round(pnl - chg, 2),
                        hold_bars   = j - position["entry_bar"],
                        entry_ts    = int(ts[position["entry_bar"]]),
                        exit_ts     = int(ts[j]),
                    )
                    trades.append(t)
                    position = None
                    break
                position["last_bar"] = j
            if position is not None:
                position["last_bar"] = min(i, len(df) - 1)
            continue

        # ── signal logic ─────────────────────────────────────────────────
        window = df.iloc[max(0, i - LOOKBACK): i + 1].copy().reset_index(drop=True)
        try:
            patterns = detector.detect_all(window)
        except Exception:
            continue

        if not patterns:
            continue

        best = patterns[0]
        if best.confidence < config.PATTERN_MIN_CONFIDENCE:
            continue

        direction = "BUY" if (best.direction or "").lower() == "bullish" else "SELL"
        trigger   = float(best.entry_price or 0)
        sl        = float(best.stop_loss  or 0)
        tgt       = float(best.target_price or 0)
        ltp       = float(close[i])

        if trigger <= 0 or sl <= 0 or tgt <= 0:
            continue

        # drift cap
        if trigger > 0:
            drift = abs(ltp - trigger) / trigger
            if drift > MAX_DRIFT:
                continue

        # direction confirmed?
        if direction == "BUY"  and ltp < trigger: continue
        if direction == "SELL" and ltp > trigger: continue

        # R:R check
        risk   = abs(ltp - sl)
        reward = abs(tgt - ltp)
        if risk <= 0 or reward / risk < RR_MIN:
            continue

        regime = detect_regime(window)["regime"]
        qty    = max(1, int(CAPITAL * SZ_PCT / ltp))

        position = {
            "entry_bar": i,
            "entry":     ltp,
            "sl":        sl,
            "tgt":       tgt,
            "direction": direction,
            "pattern":   best.pattern_name,
            "regime":    regime,
            "conf":      best.confidence,
            "qty":       qty,
            "last_bar":  i,
            "exit_price": ltp,
        }

    return trades


# ── metrics helper ────────────────────────────────────────────────────────────
def calc_metrics(trades: List[Trade]) -> dict:
    if not trades:
        return dict(n=0, wins=0, wr=0.0, gross=0.0, net=0.0, charges=0.0,
                    pf=0.0, avg_win=0.0, avg_loss=0.0, sharpe=0.0,
                    max_dd=0.0, avg_hold=0.0)
    wins   = [t for t in trades if t.net_pnl > 0]
    losses = [t for t in trades if t.net_pnl <= 0]
    gross_w = sum(t.net_pnl for t in wins)
    gross_l = abs(sum(t.net_pnl for t in losses))
    net     = sum(t.net_pnl for t in trades)
    # equity curve for Sharpe + drawdown
    curve = np.cumsum([t.net_pnl for t in trades])
    dd = 0.0
    peak = 0.0
    for v in curve:
        if v > peak: peak = v
        dd = max(dd, peak - v)
    rets = np.diff(np.concatenate([[0], curve]))
    sharpe = 0.0
    if len(rets) > 1 and np.std(rets) > 0:
        sharpe = round(float(np.mean(rets) / np.std(rets) * np.sqrt(252 * 25)), 3)
    return dict(
        n       = len(trades),
        wins    = len(wins),
        wr      = round(len(wins) / len(trades) * 100, 1),
        gross   = round(sum(t.gross_pnl for t in trades), 2),
        net     = round(net, 2),
        charges = round(sum(t.charges for t in trades), 2),
        pf      = round(gross_w / max(gross_l, 0.01), 2),
        avg_win = round(gross_w / max(len(wins), 1), 2),
        avg_loss= round(gross_l / max(len(losses), 1), 2),
        sharpe  = sharpe,
        max_dd  = round(dd, 2),
        avg_hold= round(np.mean([t.hold_bars for t in trades]), 1),
    )


# ── XLSX writer ───────────────────────────────────────────────────────────────
def write_xlsx(all_trades: List[Trade], symbol_results: dict, run_date: str):
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    wb = Workbook()
    wb.remove(wb.active)

    # ── palette ──────────────────────────────────────────────────────────────
    H1  = "16213E"; H2 = "0F3460"; GRN = "1B998B"; RED = "E74C3C"
    YEL = "F39C12"; PRP = "8E44AD"; ORG = "E67E22"
    W   = "FFFFFF"; K  = "000000"; LG  = "F2F3F4"
    LGN = "D5F5E3"; LRD = "FADBD8"; LYL = "FEF9E7"
    LPP = "E8DAEF"; LBL = "D6EAF8"

    thin = Side(style="thin", color="AEB6BF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    tkbdr= Border(left=thin, right=thin, top=thin,
                  bottom=Side(style="medium", color="2C3E50"))

    def hdr(ws, row, ncols, fill=H1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = Font(name="Arial", bold=True, size=11, color=W)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border = tkbdr

    def cell_style(ws, row, col, val, fill=None, bold=False,
                   num_fmt=None, align="center"):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name="Arial", size=10, bold=bold)
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = bdr
        if num_fmt:
            cell.number_format = num_fmt
        return cell

    def title_row(ws, row, ncols, text, fill=H1):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name="Arial", bold=True, size=14, color=W)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 26

    def col_w(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def write_table(ws, start_row, headers, rows, fill=H1, zebra=(LG, W),
                    bold_col1=True, widths=None):
        ncols = len(headers)
        for j, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=j, value=h)
        hdr(ws, start_row, ncols, fill)
        for i, row in enumerate(rows):
            bg = zebra[i % 2]
            for j, v in enumerate(row, 1):
                cell_style(ws, start_row + 1 + i, j, v,
                           fill=bg,
                           bold=(bold_col1 and j == 1),
                           align="left" if j <= 2 else "center")
        if widths:
            col_w(ws, widths)
        return start_row + 1 + len(rows)

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 1 — Summary
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("1. Summary")
    overall = calc_metrics(all_trades)
    title_row(ws, 1, 4, f"NIFTY 100 Pattern Backtest — {run_date}", H1)

    kv = [
        ("Period",         "Jan 9 – Apr 17, 2026  (65 trading days)"),
        ("Timeframe",      "15-min  (primary)"),
        ("Stocks tested",  f"{len(symbol_results)} of NIFTY 100"),
        ("Lookback",       f"{LOOKBACK} bars per pattern scan window"),
        ("Step",           f"Every {STEP} bars between scans"),
        ("Patterns",       "17 chart-pattern detectors"),
        ("Capital",        f"Rs {CAPITAL:,.0f}"),
        ("Size/trade",     f"{SZ_PCT*100:.0f}% of capital"),
        ("Min pattern conf",f"{config.PATTERN_MIN_CONFIDENCE}"),
        ("Min R:R",        f"{RR_MIN}"),
        ("Entry drift cap",f"{MAX_DRIFT*100:.1f}%"),
    ]
    write_table(ws, 3, ["Parameter", "Value"], kv,
                fill=H2, widths=[30, 40])

    r = 3 + 1 + len(kv) + 2
    title_row(ws, r, 4, "Overall Performance", GRN)
    perf_rows = [
        ("Total trades",        overall["n"]),
        ("Winners",             overall["wins"]),
        ("Losers",              overall["n"] - overall["wins"]),
        ("Win rate",            f"{overall['wr']}%"),
        ("Gross P&L",           f"Rs {overall['gross']:,.2f}"),
        ("Total charges",       f"Rs {overall['charges']:,.2f}"),
        ("Net P&L",             f"Rs {overall['net']:,.2f}"),
        ("Profit factor",       overall["pf"]),
        ("Avg win",             f"Rs {overall['avg_win']:,.2f}"),
        ("Avg loss",            f"Rs {overall['avg_loss']:,.2f}"),
        ("Sharpe ratio",        overall["sharpe"]),
        ("Max drawdown",        f"Rs {overall['max_dd']:,.2f}"),
        ("Avg holding (bars)",  overall["avg_hold"]),
    ]
    write_table(ws, r + 1, ["Metric", "Value"], perf_rows,
                fill=GRN, widths=[30, 30])

    ws.freeze_panes = "A4"

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 2 — Per Symbol
    # ─────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("2. Per Symbol")
    title_row(ws2, 1, 11, "Performance by Symbol", H2)
    sym_rows = []
    for sym, trades in sorted(symbol_results.items(),
                               key=lambda x: calc_metrics(x[1])["net"],
                               reverse=True):
        m = calc_metrics(trades)
        ticker = sym.replace("NSE:", "").replace("-EQ", "")
        sym_rows.append([
            ticker, m["n"], m["wins"], f"{m['wr']}%",
            f"Rs {m['net']:,.0f}", f"Rs {m['avg_win']:,.0f}",
            f"Rs {m['avg_loss']:,.0f}", m["pf"], m["sharpe"],
            f"Rs {m['max_dd']:,.0f}", f"{m['avg_hold']:.1f} bars",
        ])
    write_table(ws2, 3,
                ["Symbol", "Trades", "Wins", "Win%", "Net P&L",
                 "Avg Win", "Avg Loss", "P.Factor", "Sharpe",
                 "Max DD", "Avg Hold"],
                sym_rows, fill=H2, widths=[14,9,9,9,14,14,14,12,12,14,14])
    ws2.freeze_panes = "A5"

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 3 — Per Pattern
    # ─────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("3. Per Pattern")
    title_row(ws3, 1, 12, "Performance by Pattern", ORG)

    pat_map: Dict[str, List[Trade]] = defaultdict(list)
    for t in all_trades:
        pat_map[t.pattern].append(t)

    PATTERN_CATEGORIES = {
        "double_top": "REVERSAL",       "double_bottom": "REVERSAL",
        "head_shoulders_top": "REVERSAL","head_shoulders_bottom": "REVERSAL",
        "triple_top": "REVERSAL",       "triple_bottom": "REVERSAL",
        "rounding_top": "REVERSAL",     "rounding_bottom": "REVERSAL",
        "wedge": "REVERSAL",            "diamond": "REVERSAL",
        "broadening_formation": "VOLATILITY","broadening_top_bottom": "VOLATILITY",
        "ascending_triangle": "BREAKOUT","descending_triangle": "BREAKOUT",
        "symmetrical_triangle": "BREAKOUT","rectangle": "BREAKOUT",
        "cup_and_handle": "BREAKOUT",   "flag": "CONTINUATION",
        "pennant": "CONTINUATION",      "measured_move": "CONTINUATION",
    }

    pat_rows = []
    for pat_name, trades in sorted(pat_map.items(),
                                   key=lambda x: calc_metrics(x[1])["net"],
                                   reverse=True):
        m = calc_metrics(trades)
        cat = PATTERN_CATEGORIES.get(pat_name, "OTHER")
        label = pat_name.replace("_", " ").title()
        pat_rows.append([
            label, cat, m["n"], len(set(t.symbol for t in trades)),
            f"{m['wr']}%", f"Rs {m['net']:,.0f}",
            f"Rs {m['avg_win']:,.0f}", f"Rs {m['avg_loss']:,.0f}",
            m["pf"], m["sharpe"], f"Rs {m['max_dd']:,.0f}",
            f"{m['avg_hold']:.1f}",
        ])
    write_table(ws3, 3,
                ["Pattern", "Category", "Trades", "Symbols", "Win%",
                 "Net P&L", "Avg Win", "Avg Loss", "P.Factor",
                 "Sharpe", "Max DD", "Avg Hold"],
                pat_rows, fill=ORG,
                widths=[28, 16, 9, 10, 9, 14, 14, 14, 12, 12, 14, 12])
    ws3.freeze_panes = "A5"

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 4 — Regime Edge
    # ─────────────────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("4. Regime Edge")
    title_row(ws4, 1, 9, "Performance by Regime", GRN)
    reg_map: Dict[str, List[Trade]] = defaultdict(list)
    for t in all_trades:
        reg_map[t.regime or "UNKNOWN"].append(t)
    reg_rows = []
    for reg, trades in sorted(reg_map.items(),
                              key=lambda x: calc_metrics(x[1])["wr"],
                              reverse=True):
        m = calc_metrics(trades)
        reg_rows.append([
            reg, m["n"], f"{m['wr']}%",
            f"Rs {m['net']:,.0f}", m["pf"],
            f"Rs {m['avg_win']:,.0f}", f"Rs {m['avg_loss']:,.0f}",
            m["sharpe"], f"{m['avg_hold']:.1f}",
        ])
    write_table(ws4, 3,
                ["Regime", "Trades", "Win%", "Net P&L",
                 "P.Factor", "Avg Win", "Avg Loss", "Sharpe", "Avg Hold"],
                reg_rows, fill=GRN,
                widths=[22, 9, 9, 14, 12, 14, 14, 12, 12])

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 5 — Pattern × Regime heat-map (win rate)
    # ─────────────────────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("5. Pattern x Regime")
    title_row(ws5, 1, 10, "Win-Rate Heat-Map: Pattern × Regime", PRP)

    regimes = ["TRENDING_UP", "TRENDING_DOWN", "MEAN_REVERTING",
               "VOLATILE", "BREAKOUT", "CONSOLIDATION", "MOMENTUM"]
    patterns_sorted = sorted(pat_map.keys())

    # header row
    ws5.cell(row=3, column=1, value="Pattern")
    ws5.cell(row=3, column=1).font = Font(bold=True, color=W, name="Arial", size=10)
    ws5.cell(row=3, column=1).fill = PatternFill("solid", fgColor=PRP)
    ws5.cell(row=3, column=1).border = bdr
    for j, reg in enumerate(regimes, 2):
        c = ws5.cell(row=3, column=j, value=reg.replace("_", " "))
        c.font = Font(bold=True, color=W, name="Arial", size=9)
        c.fill = PatternFill("solid", fgColor=PRP)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = bdr
    ws5.column_dimensions["A"].width = 28
    for j in range(2, 2 + len(regimes)):
        ws5.column_dimensions[get_column_letter(j)].width = 16

    for i, pname in enumerate(patterns_sorted, 4):
        label = pname.replace("_", " ").title()
        c = ws5.cell(row=i, column=1, value=label)
        c.font = Font(bold=True, name="Arial", size=10)
        c.fill = PatternFill("solid", fgColor=LPP if i % 2 == 0 else W)
        c.border = bdr
        for j, reg in enumerate(regimes, 2):
            subset = [t for t in pat_map.get(pname, []) if t.regime == reg]
            if subset:
                wr = sum(1 for t in subset if t.net_pnl > 0) / len(subset) * 100
                val = f"{wr:.0f}%  (n={len(subset)})"
                clr = LGN if wr >= 55 else (LYL if wr >= 40 else LRD)
            else:
                val = "--"
                clr = LG if i % 2 == 0 else W
            c2 = ws5.cell(row=i, column=j, value=val)
            c2.font = Font(name="Arial", size=9)
            c2.fill = PatternFill("solid", fgColor=clr)
            c2.alignment = Alignment(horizontal="center")
            c2.border = bdr
    ws5.freeze_panes = "B4"

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 6 — Top 20 trades
    # ─────────────────────────────────────────────────────────────────────────
    ws6 = wb.create_sheet("6. Top 20 Trades")
    title_row(ws6, 1, 10, "Top 20 Best Individual Trades", GRN)
    top20 = sorted(all_trades, key=lambda t: t.net_pnl, reverse=True)[:20]
    _trade_rows(ws6, 3, top20, GRN)

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 7 — Worst 20 trades
    # ─────────────────────────────────────────────────────────────────────────
    ws7 = wb.create_sheet("7. Worst 20 Trades")
    title_row(ws7, 1, 10, "Worst 20 Largest Losses", RED)
    bot20 = sorted(all_trades, key=lambda t: t.net_pnl)[:20]
    _trade_rows(ws7, 3, bot20, RED)

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 8 — All Trades
    # ─────────────────────────────────────────────────────────────────────────
    ws8 = wb.create_sheet("8. All Trades")
    title_row(ws8, 1, 12, f"Full Trade Log  ({len(all_trades)} trades)", H2)
    sorted_all = sorted(all_trades, key=lambda t: t.entry_ts)
    _trade_rows(ws8, 3, sorted_all, H2, extra_cols=True)

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 9 — About / Methodology
    # ─────────────────────────────────────────────────────────────────────────
    ws9 = wb.create_sheet("9. About")
    title_row(ws9, 1, 2, "Methodology Notes", H1)
    notes = [
        ("Data source",        "SQLite cache — 15-min OHLCV (ohlcv_cache.db)"),
        ("Universe",           "NIFTY 100 (96 symbols with sufficient history)"),
        ("Date range",         "Jan 9, 2026 – Apr 17, 2026  (65 trading days)"),
        ("Bars per symbol",    "~1,625 bars of 15-min data"),
        ("Pattern engine",     "Equity Terminal  patterns/pattern_detector.py  (17 detectors)"),
        ("Entry rule",         "LTP must be at or past pattern trigger; drift ≤ 0.8%"),
        ("SL",                 "Pattern invalidation level (from pattern.stop_loss)"),
        ("Target",             "Pattern measure-rule projection (from pattern.target_price)"),
        ("R:R filter",         "Trades with R:R < 1.0 are discarded before entry"),
        ("Position size",      "10% of capital / LTP  (fixed-fraction)"),
        ("Charges",            "STT 0.025% + exchange 0.00345% + GST 18% + SEBI + Rs 20 brokerage"),
        ("Exit logic",         "SL hit / Target hit / End of data"),
        ("One trade at a time","No concurrent positions per symbol"),
        ("Regime detection",   "signals/regime_detector.py — 7 equity regimes"),
        ("Min pattern conf",   f"{config.PATTERN_MIN_CONFIDENCE}"),
        ("Lookback",           f"{LOOKBACK} bars (≈ 8 trading days)"),
        ("Scan step",          f"Every {STEP} bars when flat (every bar when in position)"),
        ("Limitations",        "No slippage model beyond drift cap; "
                               "charges are approximate; no gap-risk modelling; "
                               "single timeframe only (15-min); "
                               "patterns designed for swing / multi-day setups — "
                               "see Session Analysis for recommended intraday additions."),
    ]
    write_table(ws9, 3, ["Item", "Detail"], notes,
                fill=H1, widths=[28, 90])

    return wb


def _trade_row_data(t: Trade, extra=False) -> list:
    from datetime import datetime as _dt
    ets = _dt.fromtimestamp(t.entry_ts).strftime("%d-%b %H:%M") if t.entry_ts else "--"
    xts = _dt.fromtimestamp(t.exit_ts).strftime("%d-%b %H:%M") if t.exit_ts else "--"
    ticker = t.symbol.replace("NSE:", "").replace("-EQ", "")
    row = [
        ticker, t.direction,
        t.pattern.replace("_", " ").title(),
        t.regime, f"{t.confidence:.2f}",
        ets, xts,
        f"Rs {t.entry_price:.2f}", f"Rs {t.exit_price:.2f}",
        t.exit_reason,
        f"Rs {t.gross_pnl:,.2f}", f"Rs {t.net_pnl:,.2f}",
    ]
    if extra:
        row += [t.hold_bars, t.qty, f"Rs {t.charges:.2f}"]
    return row


def _trade_rows(ws, start_row, trades, fill, extra_cols=False):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="AEB6BF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    tkbd = Border(left=thin, right=thin, top=thin,
                  bottom=Side(style="medium", color="2C3E50"))
    W = "FFFFFF"; LGN = "D5F5E3"; LRD = "FADBD8"; LG = "F2F3F4"

    base_hdrs = ["Symbol", "Dir", "Pattern", "Regime", "Conf",
                 "Entry Time", "Exit Time", "Entry Px", "Exit Px",
                 "Reason", "Gross P&L", "Net P&L"]
    if extra_cols:
        base_hdrs += ["Hold Bars", "Qty", "Charges"]

    ncols = len(base_hdrs)
    for j, h in enumerate(base_hdrs, 1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font = Font(bold=True, color=W, name="Arial", size=11)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = tkbd

    for i, t in enumerate(trades):
        row_data = _trade_row_data(t, extra_cols)
        bg = LGN if t.net_pnl > 0 else LRD if t.net_pnl < 0 else LG
        for j, v in enumerate(row_data, 1):
            cell = ws.cell(row=start_row + 1 + i, column=j, value=v)
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center" if j > 2 else "left",
                                       vertical="center")
            cell.border = bdr

    widths = [12, 6, 26, 18, 8, 14, 14, 12, 12, 16, 14, 14]
    if extra_cols:
        widths += [10, 8, 14]
    from openpyxl.utils import get_column_letter
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = f"A{start_row + 1}"


# ── main ─────────────────────────────────────────────────────────────────────
def main():
    db_path  = config.SQLITE_DB
    det      = PatternDetector()
    chg_calc = ChargeCalculator(is_intraday=True)

    # Which NIFTY 100 symbols are in cache?
    conn = sqlite3.connect(db_path)
    c    = conn.cursor()
    c.execute("SELECT DISTINCT symbol FROM ohlcv_cache WHERE resolution='15'")
    cached = {r[0] for r in c.fetchall()}
    conn.close()

    symbols = [s for s in NIFTY_100 if s in cached]
    print(f"\nNIFTY 100 Pattern Backtest")
    print(f"{'='*60}")
    print(f"Symbols with cached data : {len(symbols)}")
    print(f"Lookback                 : {LOOKBACK} bars")
    print(f"Step                     : every {STEP} bars")
    print(f"Period                   : Jan 9 -> Apr 17, 2026\n")

    all_trades: List[Trade] = []
    symbol_results: Dict[str, List[Trade]] = {}

    t0 = time.time()
    for idx, symbol in enumerate(symbols, 1):
        df = load_symbol(symbol, db_path)
        if df is None:
            print(f"  [{idx:3}/{len(symbols)}] {symbol:<30} SKIP (insufficient bars)")
            continue
        try:
            trades = backtest_symbol(symbol, df, det, chg_calc)
            symbol_results[symbol] = trades
            all_trades.extend(trades)
            m = calc_metrics(trades)
            ticker = symbol.replace("NSE:", "").replace("-EQ", "")
            status = "OK" if m["net"] >= 0 else "LOSS"
            print(f"  [{idx:3}/{len(symbols)}] {ticker:<20} "
                  f"trades={m['n']:3}  wr={m['wr']:5.1f}%  "
                  f"net=Rs {m['net']:>9,.0f}  [{status}]")
        except Exception as e:
            print(f"  [{idx:3}/{len(symbols)}] {symbol:<30} ERROR: {e}")
            traceback.print_exc()

    elapsed = time.time() - t0
    print(f"\n{'='*60}")
    total_m = calc_metrics(all_trades)
    print(f"Total trades : {total_m['n']}")
    print(f"Win rate     : {total_m['wr']}%")
    print(f"Net P&L      : Rs {total_m['net']:,.2f}")
    print(f"Profit factor: {total_m['pf']}")
    print(f"Sharpe       : {total_m['sharpe']}")
    print(f"Elapsed      : {elapsed:.1f}s\n")

    # ── per-pattern console summary ──────────────────────────────────────────
    pat_map: Dict[str, List[Trade]] = defaultdict(list)
    for t in all_trades:
        pat_map[t.pattern].append(t)

    print(f"{'Pattern':<30} {'Trades':>6} {'Win%':>6} {'Net P&L':>12} {'PF':>6}")
    print("-" * 65)
    for pname, trades in sorted(pat_map.items(),
                                 key=lambda x: calc_metrics(x[1])["net"],
                                 reverse=True):
        m = calc_metrics(trades)
        label = pname.replace("_", " ").title()
        pnl_str = f"Rs {m['net']:>9,.0f}"
        print(f"{label:<30} {m['n']:>6}  {m['wr']:>5.1f}%  {pnl_str}  {m['pf']:>5.2f}")

    # ── write XLSX ────────────────────────────────────────────────────────────
    run_date = datetime.now().strftime("%Y%m%d_%H%M")
    out_dir  = os.path.join(ROOT, "reports")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"NIFTY100_Pattern_Backtest_{run_date}.xlsx")

    print(f"\nWriting report...")
    wb = write_xlsx(all_trades, symbol_results, run_date)
    wb.save(out_path)
    print(f"[OK] Report saved: {out_path}")


if __name__ == "__main__":
    main()
