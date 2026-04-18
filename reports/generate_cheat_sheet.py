"""
Equity Trading Terminal — Strategy Cheat Sheet Generator
========================================================
Produces a multi-tab XLSX summarising the entire pipeline so the user
can see, at a glance:

   • the 3-layer signal pipeline (GATE → QUALIFIER → TRIGGER)
   • the 7-model fusion (LightGBM + XGBoost + LSTM + TFT + ARIMA + Prophet + Pattern)
   • all 6 strategy presets and their model weights
   • all 7 market regimes and their detection rules
   • the full chart-pattern catalog with categories
   • the 6-point trade-quality scorecard with weights & thresholds
   • every risk control / cooldown / config knob
   • the complete trade lifecycle workflow
   • the Flask API surface
   • file map (module → purpose)

Run:
    python reports/generate_cheat_sheet.py
Output:
    reports/Strategy_Cheat_Sheet_Equity_Terminal.xlsx
"""
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Allow running directly from `reports/`
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config  # noqa: E402

# ──────────────────────────────────────────────────────────────────
# Palette (mirrors GEMS sheet so the two cheat sheets feel related)
# ──────────────────────────────────────────────────────────────────
HEADER_BG     = "16213E"
ACCENT_BLUE   = "0F3460"
ACCENT_GREEN  = "1B998B"
ACCENT_RED    = "E74C3C"
ACCENT_YELLOW = "F39C12"
ACCENT_PURPLE = "8E44AD"
ACCENT_ORANGE = "E67E22"
ACCENT_TEAL   = "17A589"
ACCENT_DARK   = "2C3E50"
WHITE         = "FFFFFF"
BLACK         = "000000"
LIGHT_GRAY    = "F2F3F4"
LIGHT_BLUE    = "D6EAF8"
LIGHT_GREEN   = "D5F5E3"
LIGHT_YELLOW  = "FEF9E7"
LIGHT_RED     = "FADBD8"
LIGHT_ORANGE  = "FDEBD0"
LIGHT_PURPLE  = "E8DAEF"
LIGHT_TEAL    = "D1F2EB"

thin          = Side(style="thin",   color="AEB6BF")
border_all    = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_bottom  = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="2C3E50"))


# ──────────────────────────────────────────────────────────────────
# Style helpers
# ──────────────────────────────────────────────────────────────────
def style_title(ws, row, cols, title, fill=HEADER_BG):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name="Arial", bold=True, size=16, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28


def style_subtitle(ws, row, cols, text, fill=ACCENT_DARK, color=WHITE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Arial", italic=True, size=10, color=color)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_header(ws, row, cols, fill=HEADER_BG, color=WHITE):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Arial", bold=True, size=11, color=color)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thick_bottom


def style_row(ws, row, cols, fill=None, bold_first=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Arial", size=10,
                         bold=(bold_first and c == 1))
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="left" if c <= 2 else "center",
                                   vertical="center", wrap_text=True)
        cell.border = border_all


def write_table(ws, start_row, headers, rows, header_fill=HEADER_BG,
                zebra=(LIGHT_GRAY, WHITE), col_widths=None):
    cols = len(headers)
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h)
    style_header(ws, start_row, cols, fill=header_fill)
    for i, row in enumerate(rows):
        for j, val in enumerate(row, start=1):
            ws.cell(row=start_row + 1 + i, column=j, value=val)
        style_row(ws, start_row + 1 + i, cols,
                  fill=zebra[i % 2], bold_first=True)
    if col_widths:
        for j, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    return start_row + 1 + len(rows)


# ──────────────────────────────────────────────────────────────────
# Workbook
# ──────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)


# ──────────────────────────────────────────────────────────────────
# 1.  Overview
# ──────────────────────────────────────────────────────────────────
def sheet_overview():
    ws = wb.create_sheet("1. Overview")
    style_title(ws, 1, 4, "EQUITY TRADING TERMINAL — Cheat Sheet")
    style_subtitle(ws, 2, 4,
        "NSE cash-market intraday • Fyers API v3 • 7-model fusion • 6-point quality gate")

    rows = [
        ("Project",          "Equity Trading Terminal (NSE/BSE cash market)"),
        ("Asset class",      "Listed equities, intraday + positional swing"),
        ("Broker / data",    "Fyers API v3 (OAuth2)"),
        ("Default universe", config.DEFAULT_UNIVERSE + " (200 large-caps)"),
        ("Primary timeframe","15-minute (multi-TF context: 5 / 15 / 60)"),
        ("History cached",   f"{config.CANDLE_HISTORY_DAYS} days (SQLite)"),
        ("Poll interval",    f"{config.POLL_INTERVAL} s scan loop"),
        ("Max stocks/cycle", f"{config.MAX_STOCKS_PER_CYCLE}"),
        ("Web port",         f"{config.PORT} (Flask + SocketIO)"),
        ("Capital",          f"₹ {config.TOTAL_TRADING_CAPITAL:,.0f}"),
        ("Max concurrent",   f"{config.MAX_CONCURRENT_POSITIONS} positions"),
        ("Max per stock",    f"{config.MAX_CAPITAL_PER_STOCK*100:.0f}% of capital"),
        ("Daily loss cap",   f"₹ {config.MAX_LOSS_PER_DAY:,.0f}"),
        ("Quality gate",     f"AUTO requires grade ≥ C ({config.TRADE_QUALITY_MIN_SCORE} score)"),
        ("Patterns detected","17 chart patterns across 4 categories"),
        ("ML models",        "7 (LightGBM, XGBoost, LSTM, TFT, ARIMA, Prophet, Meta)"),
        ("Strategy presets", "6 named weight schemes"),
        ("Regimes",          "7 (Trending Up/Dn, Mean-Rev, Volatile, Breakout, Consolidation, Momentum)"),
    ]
    write_table(ws, 4, ["Property", "Value"], rows,
                col_widths=[26, 60])

    # Pipeline diagram
    r = 4 + 1 + len(rows) + 2
    style_title(ws, r, 4, "3-Layer Signal Pipeline", fill=ACCENT_BLUE)
    pipeline = [
        ("LAYER 1 — GATE",
         "Market hours · Holidays · Circuit limits",
         "Hard block, no softening",
         "entry_gate.py"),
        ("LAYER 2 — QUALIFIER",
         "ADX ≥ 10 · Volume > 0 · No 8% gap last 3 bars · A/D ≥ 0.15",
         "Hard block on dead/extreme markets",
         "signal_engine._qualifier_check"),
        ("LAYER 3 — TRIGGER",
         "Patterns → Features → 7-model predict → Fusion → Cooldown → R:R "
         "→ Entry-gate → Pattern dedup → Trade-quality → Trade engine",
         "Soft penalties, regime-weighted",
         "signal_engine._trigger"),
    ]
    write_table(ws, r + 1,
                ["Layer", "Checks", "Behaviour", "Source"], pipeline,
                header_fill=ACCENT_BLUE, col_widths=[20, 60, 32, 32])

sheet_overview()


# ──────────────────────────────────────────────────────────────────
# 2.  7-Model Fusion
# ──────────────────────────────────────────────────────────────────
def sheet_models():
    ws = wb.create_sheet("2. Models")
    style_title(ws, 1, 5, "The 7-Model Ensemble", fill=ACCENT_PURPLE)
    style_subtitle(ws, 2, 5,
        "Each model emits a probability/direction; SignalFusion blends them "
        "with regime-aware weights (or static fallback).")

    rows = [
        ("Pattern detector", "Rule-based",
         "17 chart patterns + swing/trendline engine",
         f"{config.FUSION_FALLBACK_WEIGHTS['pattern']*100:.0f}%",
         "patterns/pattern_detector.py"),
        ("LightGBM",         "Gradient boosting",
         "Tabular features (RSI/MACD/BB/ATR/ADX/VWAP/OBV + lags)",
         f"{config.FUSION_FALLBACK_WEIGHTS['lgbm']*100:.0f}%",
         "models/lgbm_model.py"),
        ("XGBoost",          "Gradient boosting",
         "Same features, different bias",
         f"{config.FUSION_FALLBACK_WEIGHTS['xgb']*100:.0f}%",
         "models/xgboost_model.py"),
        ("LSTM",             "Recurrent NN",
         f"Sequence of last {config.ML_LOOKBACK_TIMESTEPS} bars (OHLCV + indicators)",
         f"{config.FUSION_FALLBACK_WEIGHTS['lstm']*100:.0f}%",
         "models/lstm_model.py"),
        ("TFT",              "Temporal Fusion Transformer",
         "Attention over sequence + static covariates",
         f"{config.FUSION_FALLBACK_WEIGHTS['tft']*100:.0f}%",
         "models/tft_model.py"),
        ("ARIMA",            "Statistical",
         "Trend direction (UP / DOWN / FLAT)",
         f"{config.FUSION_FALLBACK_WEIGHTS['arima']*100:.0f}%",
         "models/arima_model.py"),
        ("Prophet",          "Statistical",
         "Daily / weekly seasonality (rarely used)",
         f"{config.FUSION_FALLBACK_WEIGHTS['prophet']*100:.0f}%",
         "models/prophet_model.py"),
        ("FII / DII context","Enricher",
         "Net institutional flows tilt the score",
         f"{config.FUSION_FALLBACK_WEIGHTS['fii']*100:.0f}%",
         "core/market_data_enricher.py"),
        ("OI context",       "Enricher",
         "PCR & open-interest tilt the score",
         f"{config.FUSION_FALLBACK_WEIGHTS['oi']*100:.0f}%",
         "core/market_data_enricher.py"),
        ("Meta-learner",     "Logistic regression",
         "Optional layer that re-blends base predictions",
         "stack",
         "models/meta_learner.py"),
    ]
    write_table(ws, 4,
                ["Model", "Family", "What it consumes",
                 "Default weight", "File"],
                rows, header_fill=ACCENT_PURPLE,
                col_widths=[18, 22, 56, 14, 30])

    # ML thresholds
    r = 4 + 1 + len(rows) + 2
    style_title(ws, r, 5, "Training & Threshold Settings", fill=ACCENT_PURPLE)
    thresholds = [
        ("ML_LOOKBACK_TIMESTEPS", config.ML_LOOKBACK_TIMESTEPS,
         "Sequence length fed to LSTM / TFT"),
        ("ML_PREDICTION_HORIZON", config.ML_PREDICTION_HORIZON,
         "Bars ahead the model labels (≈ 1.5 h on 15-m)"),
        ("ML_TRAIN_DAYS",         config.ML_TRAIN_DAYS,
         "Rolling training window in days"),
        ("ML_MIN_ACCURACY",       config.ML_MIN_ACCURACY,
         "Force retrain if walk-forward accuracy drops below"),
        ("ML_RETRAIN_DAY",        config.ML_RETRAIN_DAY,
         "Scheduled weekly retrain day"),
        ("PATTERN_MIN_CONFIDENCE",config.PATTERN_MIN_CONFIDENCE,
         "Min pattern conf to invoke ML predict"),
    ]
    write_table(ws, r + 1, ["Setting", "Value", "Meaning"], thresholds,
                header_fill=ACCENT_PURPLE, col_widths=[28, 14, 60])

sheet_models()


# ──────────────────────────────────────────────────────────────────
# 3.  Strategy Presets
# ──────────────────────────────────────────────────────────────────
def sheet_strategies():
    ws = wb.create_sheet("3. Strategy Presets")
    style_title(ws, 1, 11, "6 Strategy Presets — Model Weight Mix", fill=ACCENT_GREEN)
    style_subtitle(ws, 2, 11,
        "Selected via UI dropdown or auto-adapted by AdaptiveBot. "
        "Each preset re-weights the 7-model fusion.")

    from signals.strategy_presets import PRESETS

    headers = ["Preset", "Pattern", "LGBM", "XGB", "LSTM", "TFT",
               "ARIMA", "Prophet", "FII", "OI", "Min Conf"]
    rows = []
    for name, p in PRESETS.items():
        w = p.weights
        rows.append([
            f"{name}\n({p.description})",
            f"{w['pattern']*100:.0f}%",
            f"{w['lgbm']*100:.0f}%",
            f"{w['xgb']*100:.0f}%",
            f"{w['lstm']*100:.0f}%",
            f"{w['tft']*100:.0f}%",
            f"{w['arima']*100:.0f}%",
            f"{w['prophet']*100:.0f}%",
            f"{w['fii']*100:.0f}%",
            f"{w['oi']*100:.0f}%",
            f"{p.min_confidence:.0f}",
        ])
    write_table(ws, 4, headers, rows,
                header_fill=ACCENT_GREEN,
                zebra=(LIGHT_GREEN, WHITE),
                col_widths=[34, 9, 9, 9, 9, 9, 9, 9, 9, 9, 10])

    # Required conditions
    r = 4 + 1 + len(rows) + 2
    style_title(ws, r, 11, "Required Conditions per Preset", fill=ACCENT_GREEN)
    cond_rows = []
    for name, p in PRESETS.items():
        if p.required_conditions:
            for k, v in p.required_conditions.items():
                cond_rows.append([name, k, v])
        else:
            cond_rows.append([name, "(none)", "—"])
    write_table(ws, r + 1, ["Preset", "Required indicator", "Min value"],
                cond_rows, header_fill=ACCENT_GREEN,
                zebra=(LIGHT_GREEN, WHITE), col_widths=[28, 28, 14])

sheet_strategies()


# ──────────────────────────────────────────────────────────────────
# 4.  Pattern Catalog
# ──────────────────────────────────────────────────────────────────
def sheet_patterns():
    ws = wb.create_sheet("4. Patterns")
    style_title(ws, 1, 4, "Chart Pattern Catalog (17 detectors)", fill=ACCENT_ORANGE)
    style_subtitle(ws, 2, 4,
        "Each detector returns a confidence ∈ [0,1] and trigger price. "
        "Category is used by trade_quality.py → regime fit check.")

    rows = [
        ("Double Top",            "REVERSAL",     "reversal_patterns.py",
         "Two equal highs with neckline support; bearish on neckline break"),
        ("Double Bottom",         "REVERSAL",     "reversal_patterns.py",
         "Two equal lows with neckline resistance; bullish on neckline break"),
        ("Head & Shoulders Top",  "REVERSAL",     "reversal_patterns.py",
         "Higher-high middle bracketed by lower highs; bearish"),
        ("Head & Shoulders Bottom","REVERSAL",    "reversal_patterns.py",
         "Inverse H&S; bullish reversal off downtrend"),
        ("Triple Top",            "REVERSAL",     "reversal_patterns.py",
         "Three equal-resistance failures; bearish"),
        ("Triple Bottom",         "REVERSAL",     "reversal_patterns.py",
         "Three equal-support tests; bullish"),
        ("Rounding Top",          "REVERSAL",     "reversal_patterns.py",
         "Curved distribution top; slow bearish reversal"),
        ("Rounding Bottom",       "REVERSAL",     "reversal_patterns.py",
         "Saucer accumulation; slow bullish reversal"),
        ("Wedge (rising/falling)","REVERSAL",     "volatility_patterns.py",
         "Converging trendlines, both sloping same direction"),
        ("Diamond",               "REVERSAL",     "volatility_patterns.py",
         "Broadening then narrowing, often distribution top"),
        ("Broadening Formation",  "VOLATILITY",   "volatility_patterns.py",
         "Diverging highs/lows; high-vol regime"),
        ("Broadening Top/Bottom", "VOLATILITY",   "volatility_patterns.py",
         "Asymmetric broadening with directional bias"),
        ("Ascending Triangle",    "CONTINUATION", "breakout_patterns.py",
         "Flat top / rising lows; bullish on top break"),
        ("Descending Triangle",   "CONTINUATION", "breakout_patterns.py",
         "Flat bottom / falling highs; bearish on bottom break"),
        ("Symmetrical Triangle",  "CONTINUATION", "breakout_patterns.py",
         "Converging trendlines; trade in breakout direction"),
        ("Rectangle",             "CONTINUATION", "breakout_patterns.py",
         "Horizontal range; trade the breakout"),
        ("Cup & Handle",          "CONTINUATION", "breakout_patterns.py",
         "U-base + small pullback; bullish continuation"),
        ("Flag",                  "CONTINUATION", "continuation_patterns.py",
         "Sharp move then small parallel pullback"),
        ("Pennant",               "CONTINUATION", "continuation_patterns.py",
         "Sharp move then small symmetric triangle"),
        ("Measured Move",         "CONTINUATION", "continuation_patterns.py",
         "ABC swing whose C-leg targets length of A-leg"),
    ]
    write_table(ws, 4,
                ["Pattern", "Category", "Source file", "What it means"],
                rows, header_fill=ACCENT_ORANGE,
                zebra=(LIGHT_ORANGE, WHITE),
                col_widths=[26, 18, 28, 65])

    # Pattern engine config
    r = 4 + 1 + len(rows) + 2
    style_title(ws, r, 4, "Pattern Engine Settings", fill=ACCENT_ORANGE)
    pattern_cfg = [
        ("SWING_LEFT_BARS",          config.SWING_LEFT_BARS,
         "Bars left of pivot to confirm swing high/low"),
        ("SWING_RIGHT_BARS",         config.SWING_RIGHT_BARS,
         "Bars right of pivot to confirm swing"),
        ("SWING_MIN_AMPLITUDE_PCT",  config.SWING_MIN_AMPLITUDE_PCT,
         "Minimum swing size to count (1% = 0.01)"),
        ("PATTERN_LOOKBACK_BARS",    config.PATTERN_LOOKBACK_BARS,
         "Bars to scan for pattern formation"),
        ("PATTERN_MIN_CONFIDENCE",   config.PATTERN_MIN_CONFIDENCE,
         "Below this, ML pipeline is skipped entirely"),
        ("TRENDLINE_MIN_R_SQUARED",  config.TRENDLINE_MIN_R_SQUARED,
         "Min R² for a trendline to be considered valid"),
        ("TRENDLINE_MIN_TOUCHES",    config.TRENDLINE_MIN_TOUCHES,
         "Min number of pivots a trendline must touch"),
    ]
    write_table(ws, r + 1, ["Setting", "Value", "Meaning"], pattern_cfg,
                header_fill=ACCENT_ORANGE,
                zebra=(LIGHT_ORANGE, WHITE), col_widths=[28, 14, 65])

sheet_patterns()


# ──────────────────────────────────────────────────────────────────
# 5.  Regimes
# ──────────────────────────────────────────────────────────────────
def sheet_regimes():
    ws = wb.create_sheet("5. Regimes")
    style_title(ws, 1, 4, "7 Equity Market Regimes", fill=ACCENT_TEAL)
    style_subtitle(ws, 2, 4,
        "Detected per-symbol from ADX / ATR / slope / pivot density. "
        "SignalFusion uses the regime to re-weight the ensemble.")

    rows = [
        ("TRENDING_UP",     f"ADX ≥ {config.REGIME_ADX_TREND_MIN} & positive slope",
         "Continuation patterns BUY-side; reversals fade",
         "Favours LSTM / TFT / pattern continuation weights"),
        ("TRENDING_DOWN",   f"ADX ≥ {config.REGIME_ADX_TREND_MIN} & negative slope",
         "Continuation patterns SELL-side; reversals fade",
         "Favours LSTM / TFT / pattern continuation weights"),
        ("MEAN_REVERTING",  f"ADX ≤ {config.REGIME_ADX_MEANREV_MAX} & oscillating",
         "Reversal patterns at extremes",
         "Favours LightGBM / XGB tabular models"),
        ("VOLATILE",        f"ATR > {config.REGIME_ATR_VOLATILE_MULT}× average",
         "Wide ranges, broadening patterns",
         "Penalty applied; sizes shrunk"),
        ("BREAKOUT",        "Range compression then expansion",
         "Triangle / rectangle / pennant breakouts",
         "Boosts pattern weight"),
        ("CONSOLIDATION",   f"ADX ≤ {config.REGIME_ADX_CONSOLIDATION_MAX}",
         "Low-vol drift; wait for breakout",
         "Default fallback regime"),
        ("MOMENTUM",        "Persistent directional thrust w/ strong volume",
         "Continuation in same direction",
         "Boosts XGB + LSTM weights"),
    ]
    write_table(ws, 4,
                ["Regime", "Detection rule", "Pattern fit", "Effect on fusion"],
                rows, header_fill=ACCENT_TEAL,
                zebra=(LIGHT_TEAL, WHITE),
                col_widths=[20, 40, 38, 42])

sheet_regimes()


# ──────────────────────────────────────────────────────────────────
# 6.  6-Point Trade Quality Scorecard
# ──────────────────────────────────────────────────────────────────
def sheet_quality():
    ws = wb.create_sheet("6. Trade Quality")
    style_title(ws, 1, 5, "6-Point Trade Quality Scorecard (A → F)",
                fill=ACCENT_RED)
    style_subtitle(ws, 2, 5,
        "Runs after fusion produces a candidate, BEFORE the trade engine. "
        "Hard AUTO gate in trade_engine.process_signal.")

    rows = [
        ("1. Breakout freshness",  "freshness", "20%",
         f"|entry − trigger| / trigger ≤ {config.TRADE_QUALITY_MAX_DRIFT_PCT}%",
         "Stale breakouts eat all the R:R"),
        ("2. Risk size",           "risk_size", "20%",
         f"|entry − SL| / entry ≥ {config.TRADE_QUALITY_MIN_RISK_PCT}%",
         "Prevents artificially tight SL inflating R:R"),
        ("3. R:R sanity",          "rr",        "15%",
         f"{config.TRADE_QUALITY_MIN_RR}× ≤ R:R ≤ {config.TRADE_QUALITY_MAX_RR}×",
         "Above 10× is almost always a tight-SL artefact"),
        ("4. SL on correct side",  "sl_side",   "20%",
         "BUY: SL ≤ min(swing lows); SELL: SL ≥ max(swing highs)",
         "SL must sit beyond pattern invalidation"),
        ("5. Regime fit",          "regime",    "10%",
         "Reversal in MEAN_REV / Continuation in TRENDING-same-direction",
         "Pattern category × regime taxonomy"),
        ("6. ML alignment",        "ml",        "15%",
         "≥ 2 of 5 base models (LGB+XGB+LSTM+TFT+ARIMA) agree",
         "0.55 BUY / 0.45 SELL probability thresholds"),
    ]
    write_table(ws, 4,
                ["Check", "Key", "Weight", "Pass condition", "Why"],
                rows, header_fill=ACCENT_RED,
                zebra=(LIGHT_RED, WHITE),
                col_widths=[26, 14, 10, 50, 50])

    # Grade thresholds
    r = 4 + 1 + len(rows) + 2
    style_title(ws, r, 5, "Grade Thresholds", fill=ACCENT_RED)
    grades = [
        ("A", "≥ 0.85", "All checks essentially perfect"),
        ("B", "0.70 – 0.84", "Solid; minor weakness"),
        ("C", "0.55 – 0.69", "Marginal; AUTO gate cut-off at 0.60"),
        ("D", "0.40 – 0.54", "Weak — blocked in AUTO, queued in MANUAL with ⚠"),
        ("F", "< 0.40", "Reject — multiple critical fails"),
    ]
    write_table(ws, r + 1, ["Grade", "Score range", "Meaning"],
                grades, header_fill=ACCENT_RED,
                zebra=(LIGHT_RED, WHITE), col_widths=[12, 18, 60])

    # Knobs
    r += 1 + len(grades) + 2
    style_title(ws, r, 5, "Quality Scorecard Config Knobs", fill=ACCENT_RED)
    knobs = [
        ("TRADE_QUALITY_ENABLED",       config.TRADE_QUALITY_ENABLED,
         "Master switch"),
        ("TRADE_QUALITY_MIN_SCORE",     config.TRADE_QUALITY_MIN_SCORE,
         "AUTO-mode hard gate (≥ grade C)"),
        ("TRADE_QUALITY_MIN_RR",        config.TRADE_QUALITY_MIN_RR,
         "Below → fail R:R check"),
        ("TRADE_QUALITY_MAX_RR",        config.TRADE_QUALITY_MAX_RR,
         "Above → fail R:R (likely SL artefact)"),
        ("TRADE_QUALITY_MIN_RISK_PCT",  config.TRADE_QUALITY_MIN_RISK_PCT,
         "Below → fail risk size"),
        ("TRADE_QUALITY_MAX_DRIFT_PCT", config.TRADE_QUALITY_MAX_DRIFT_PCT,
         "Above → fail freshness"),
    ]
    write_table(ws, r + 1, ["Setting", "Value", "Meaning"], knobs,
                header_fill=ACCENT_RED,
                zebra=(LIGHT_RED, WHITE), col_widths=[32, 14, 60])

sheet_quality()


# ──────────────────────────────────────────────────────────────────
# 7.  Risk Controls & Cooldowns
# ──────────────────────────────────────────────────────────────────
def sheet_risk():
    ws = wb.create_sheet("7. Risk & Cooldowns")
    style_title(ws, 1, 4, "Risk Controls, Cooldowns & Dedup", fill=ACCENT_RED)
    style_subtitle(ws, 2, 4,
        "Layered defenses: signal-level, pattern-level, position-level, "
        "portfolio-level, daily-level.")

    rows = [
        ("Pattern-instance dedup",
         "(symbol, pattern_name, round(trigger,1), direction)",
         "Same neckline / breakdown fires only ONCE per IST day",
         "signals/signal_engine.py"),
        ("Signal cooldown",
         f"{config.SIGNAL_ENTRY_COOLDOWN} s between any two signals on same symbol",
         "Hard floor independent of regime",
         "signals/signal_engine.py"),
        ("Post-trade cooldown",
         f"Base {config.POST_TRADE_COOLDOWN_SEC} s; ×2 per consecutive loss; cap 4 h",
         "Stops 'close → re-enter same pattern' loop",
         "trading/trade_engine.py"),
        ("Duplicate-symbol guard",
         "Never two ACTIVE/PENDING positions on same stock",
         "Hard block in process_signal()",
         "trading/trade_engine.py"),
        ("AUTO quality gate",
         f"signal.quality_passed required (score ≥ {config.TRADE_QUALITY_MIN_SCORE})",
         "Manual mode still queues with ⚠ badge",
         "trading/trade_engine.py"),
        ("Daily loss limit",
         f"₹ {config.MAX_LOSS_PER_DAY:,.0f}",
         "All new entries blocked once breached",
         "trading/trade_engine.py"),
        ("Min R:R",
         f"{config.MIN_RISK_REWARD}",
         "Reward / risk floor at signal level",
         "signals/signal_engine.py"),
        ("Entry-drift cap",
         f"{config.MAX_ENTRY_DRIFT_PCT}% (LTP vs trigger)",
         "Suppresses stale breakouts (phantom-fill bug fix)",
         "signals/signal_engine._compute_levels"),
        ("Trailing stop",
         f"Activate at {config.TRAILING_STOP_ACTIVATION}× initial risk; "
         f"trail by {config.TRAILING_STOP_ATR_MULT}× ATR",
         "Locks in profit once R-multiple reached",
         "trading/trade_engine._update_trailing_stop"),
        ("Position size",
         f"≤ {config.MAX_PORTFOLIO_RISK*100:.0f}% portfolio risk · "
         f"≤ {config.MAX_CAPITAL_PER_STOCK*100:.0f}% capital per stock",
         "Risk-based sizing from SL distance",
         "trading/risk_manager.py"),
        ("Rate limit",
         f"{config.MAX_REQ_PER_SEC}/sec, {config.MAX_REQ_PER_MIN}/min, "
         f"batch {config.BATCH_SIZE}, delay {config.BATCH_DELAY}s",
         "Buffer below Fyers' 10/sec, 200/min cap",
         "core/rate_limiter.py"),
        ("No-signal windows",
         f"First {config.NO_SIGNAL_FIRST_MIN} min after open · "
         f"last {config.NO_SIGNAL_LAST_MIN} min before close",
         "Avoids opening-volatility & MIS auto-square-off zone",
         "config + entry_gate"),
        ("Holiday calendar",
         f"{len(config.NSE_HOLIDAYS_2026)} NSE holidays for 2026",
         "Hard GATE block on holidays",
         "config.NSE_HOLIDAYS_2026"),
    ]
    write_table(ws, 4,
                ["Control", "Setting", "Effect", "Source"],
                rows, header_fill=ACCENT_RED,
                zebra=(LIGHT_RED, WHITE),
                col_widths=[28, 50, 50, 36])

sheet_risk()


# ──────────────────────────────────────────────────────────────────
# 8.  Workflow (trade lifecycle)
# ──────────────────────────────────────────────────────────────────
def sheet_workflow():
    ws = wb.create_sheet("8. Workflow")
    style_title(ws, 1, 3, "End-to-End Trade Lifecycle", fill=ACCENT_BLUE)
    style_subtitle(ws, 2, 3,
        "Each row is one step in the loop running every "
        f"{config.POLL_INTERVAL} s.")

    steps = [
        ("01", "Universe rotation",
         "WatchlistManager picks next batch of symbols (NIFTY_200 ÷ batch size)"),
        ("02", "Quote + history fetch",
         "fyers_manager.get_batch_quotes / get_history (rate-limited, SQLite cached)"),
        ("03", "Indicator pipeline",
         "data_engine adds RSI/MACD/BB/ATR/ADX/VWAP/OBV + lags"),
        ("04", "GATE",
         "Market hours · holiday · circuit-limit checks (entry_gate)"),
        ("05", "QUALIFIER",
         "ADX ≥ 10 · volume > 0 · no 8% gap · A/D ≥ 0.15"),
        ("06", "Pattern detection",
         "17 detectors run on swings; ranked by confidence"),
        ("07", "Feature build",
         "feature_pipeline.build_features(df, patterns, enricher)"),
        ("08", "Regime detection",
         "regime_detector.detect_regime → one of 7 regimes"),
        ("09", "ML predict",
         "model_manager.predict (LGB+XGB+LSTM+TFT+ARIMA+Prophet+Meta)"),
        ("10", "Signal fusion",
         "signal_fusion.compute → confidence + direction + strength"),
        ("11", "Threshold + cooldown",
         f"conf ≥ {config.SIGNAL_WEAK_THRESHOLD} · cooldown ≥ {config.SIGNAL_ENTRY_COOLDOWN}s"),
        ("12", "Entry / SL / Target",
         "_compute_levels(): trigger confirmed + drift cap + R:R ≥ "
         f"{config.MIN_RISK_REWARD}"),
        ("13", "Entry-gate validate",
         "Final pre-execution sanity (size, side, exposure)"),
        ("14", "Pattern-instance dedup",
         "Skip if (symbol, pattern, round(trigger,1), dir) already fired today"),
        ("15", "Trade-quality scoring",
         "6-point scorecard → A/B/C/D/F + fail_reasons[] attached to signal"),
        ("16", "Build TradeSignal",
         "Includes pattern_trigger, swings, all probs, quality_grade"),
        ("17", "Stamp dedup + cooldown",
         "Add tuple to daily set; record signal timestamp"),
        ("18", "Trade engine intake",
         "process_signal: duplicate-guard → cooldown → daily-loss → "
         "QUALITY-GATE (AUTO only)"),
        ("19", "Position size",
         "risk_manager.calculate_position_size(entry, SL)"),
        ("20", "Execute",
         "AUTO → place_intraday_order (LIMIT + MPP buffer);  "
         "MANUAL → status='PENDING' for user confirm"),
        ("21", "Live monitoring",
         "update_positions(quotes): SL / target / trailing-stop checks"),
        ("22", "Exit",
         "STOP_LOSS / TRAILING_STOP / TARGET / MANUAL_CLOSE / EOD square-off"),
        ("23", "Post-trade cooldown",
         "POST_TRADE_COOLDOWN_SEC × 2^(consecutive_losses), capped 4 h"),
        ("24", "Telegram alert",
         "Signal & exit notifications sent to configured chat"),
        ("25", "Persist + report",
         "trades.json autosaved (merge-on-save); XLSX report on demand"),
        ("26", "Bayesian update",
         "bayesian_selector.update(strategy, sector, regime, won) tightens priors"),
    ]
    write_table(ws, 4, ["#", "Step", "What happens"], steps,
                header_fill=ACCENT_BLUE,
                zebra=(LIGHT_BLUE, WHITE),
                col_widths=[6, 26, 90])

sheet_workflow()


# ──────────────────────────────────────────────────────────────────
# 9.  API Endpoints
# ──────────────────────────────────────────────────────────────────
def sheet_api():
    ws = wb.create_sheet("9. API")
    style_title(ws, 1, 3, "Flask API Surface", fill=ACCENT_DARK)
    style_subtitle(ws, 2, 3,
        f"All routes served on http://localhost:{config.PORT}.  "
        "SocketIO room 'dashboard' pushes live updates.")

    rows = [
        # Auth
        ("GET",  "/",                              "Single-page dashboard"),
        ("GET",  "/auth",                          "Fyers OAuth2 init (UI)"),
        ("POST", "/auth/callback",                 "OAuth code → access token"),
        ("POST", "/auth/logout",                   "Drop token"),
        ("POST", "/api/credentials",               "Save FYERS_APP_ID + FYERS_SECRET"),
        ("GET",  "/api/auth/status",               "Is Fyers session live?"),
        # Status / config
        ("GET",  "/api/status",                    "Engine running, strategy, regime, etc."),
        ("GET",  "/api/regime",                    "Current market regime"),
        ("GET",  "/api/strategies",                "List strategy presets"),
        ("POST", "/api/start",                     "Start scanner loop"),
        ("POST", "/api/stop",                      "Stop scanner loop"),
        ("POST", "/api/set_universe",              "NIFTY_50 / 100 / 200 / CUSTOM"),
        ("POST", "/api/set_strategy",              "Switch active preset"),
        ("POST", "/api/toggle_auto_execute",       "Toggle AUTO ↔ MANUAL"),
        ("POST", "/api/toggle_auto_adapt",         "Toggle AdaptiveBot"),
        ("POST", "/api/config/capital",            "Set total capital + caps"),
        ("GET",  "/api/config/capital/get",        "Read capital config"),
        ("POST", "/api/config/order_settings",     "MPP %, staleness sec"),
        # Data
        ("GET",  "/api/indices",                   "NIFTY/BANKNIFTY/sector indices"),
        ("GET",  "/api/scanner",                   "Latest scanner table snapshot"),
        ("GET",  "/api/portfolio",                 "Active+pending positions"),
        ("GET",  "/api/portfolio/stats",           "All-time + today metrics"),
        ("GET",  "/api/positions",                 "Active positions only"),
        ("POST", "/api/positions/cleanup",         "Dedup duplicate positions"),
        ("POST", "/api/confirm_trade/<pos_id>",    "MANUAL mode confirm"),
        ("POST", "/api/cancel_trade/<pos_id>",     "Cancel pending or close active"),
        ("GET",  "/api/universe_symbols",          "Current universe list"),
        ("GET",  "/api/chart_data",                "OHLCV for Lightweight Charts"),
        ("GET",  "/api/pattern_overlay",           "Pattern markers + lines"),
        # Models
        ("GET",  "/api/models",                    "Trained-model inventory"),
        ("POST", "/api/train",                     "Train one symbol"),
        ("GET",  "/api/train/queue",               "Background-training progress"),
        ("POST", "/api/train/all",                 "Train every symbol in universe"),
        # Reports / notify / webhook
        ("GET",  "/api/report/download",           "Build & download trade XLSX"),
        ("POST", "/api/telegram/config",           "Save bot_token + chat_id"),
        ("GET",  "/api/telegram/status",           "Is Telegram alerting wired up"),
        ("POST", "/api/telegram/test",             "Send test message"),
        ("POST", "/api/webhook",                   "TradingView → trade webhook"),
    ]
    write_table(ws, 4, ["Method", "Route", "Purpose"], rows,
                header_fill=ACCENT_DARK,
                zebra=(LIGHT_GRAY, WHITE),
                col_widths=[10, 38, 60])

sheet_api()


# ──────────────────────────────────────────────────────────────────
# 10. File Map
# ──────────────────────────────────────────────────────────────────
def sheet_files():
    ws = wb.create_sheet("10. File Map")
    style_title(ws, 1, 2, "Module → Purpose Reference", fill=ACCENT_BLUE)

    rows = [
        ("app.py",                                "Flask + SocketIO server, scanner loop, dispatch"),
        ("config.py",                             "All knobs: creds, thresholds, charges, paths"),
        ("requirements.txt",                      "Python dependencies"),
        ("templates/index.html",                  "Single-page dashboard (Tailwind + Lightweight Charts)"),
        ("static/",                               "Bundled JS + icons"),
        ("core/fyers_manager.py",                 "OAuth2, quotes, history, batch quotes, orders"),
        ("core/data_engine.py",                   "OHLCV cache + indicator computation (SQLite)"),
        ("core/market_data_enricher.py",          "PCR, FII/DII, A/D, delivery%, sector tagging"),
        ("core/stock_universe.py",                "NIFTY_50 / 100 / 200 + sector map"),
        ("core/watchlist_manager.py",             "Active scan rotation + per-symbol prioritisation"),
        ("core/rate_limiter.py",                  "8/sec, 150/min token bucket"),
        ("patterns/swing_detector.py",            "Fractal pivot detection"),
        ("patterns/trendline_engine.py",          "Linear-regression trendlines (R²-validated)"),
        ("patterns/reversal_patterns.py",         "H&S, Double/Triple top/bottom, Rounding"),
        ("patterns/continuation_patterns.py",     "Flag, Pennant, Measured Move"),
        ("patterns/breakout_patterns.py",         "Triangles, Rectangle, Cup & Handle"),
        ("patterns/volatility_patterns.py",       "Wedge, Diamond, Broadening"),
        ("patterns/pattern_validator.py",         "Volume / breakout confirmation"),
        ("patterns/multi_tf_validator.py",        "Cross-timeframe (5/15/60) agreement"),
        ("patterns/pattern_detector.py",          "Orchestrator → ranked PatternResult list"),
        ("features/feature_pipeline.py",          "Tabular + sequence features for ML"),
        ("models/lgbm_model.py",                  "LightGBM tabular classifier"),
        ("models/xgboost_model.py",               "XGBoost tabular classifier"),
        ("models/lstm_model.py",                  "LSTM sequence model"),
        ("models/tft_model.py",                   "Temporal Fusion Transformer"),
        ("models/arima_model.py",                 "ARIMA trend"),
        ("models/prophet_model.py",               "Prophet seasonality"),
        ("models/meta_learner.py",                "Logistic regression stack"),
        ("models/model_trainer.py",               "Walk-forward training, parallel"),
        ("models/model_manager.py",               "Per-symbol model load + predict"),
        ("signals/signal_engine.py",              "3-layer pipeline (THE entry point)"),
        ("signals/signal_fusion.py",              "Regime-weighted blend + soft penalties"),
        ("signals/regime_detector.py",            "7 equity-market regimes"),
        ("signals/entry_gate.py",                 "Final pre-execution validation"),
        ("signals/strategy_presets.py",           "6 named weight schemes"),
        ("signals/trade_quality.py",              "6-point pre-trade scorecard (A–F)"),
        ("trading/trade_engine.py",               "Position lifecycle + AUTO quality gate"),
        ("trading/risk_manager.py",               "Position sizing & exposure caps"),
        ("trading/portfolio_allocator.py",        "Cross-portfolio signal ranking"),
        ("trading/charge_calculator.py",          "STT / GST / brokerage / SEBI"),
        ("trading/telegram_notifier.py",          "Signal + exit alerts"),
        ("trading/webhook_executor.py",           "TradingView → trade bridge"),
        ("trading/generate_report.py",            "Daily XLSX trade report"),
        ("adaptive/adaptive_bot.py",              "Auto-switch preset by regime"),
        ("adaptive/bayesian_selector.py",         "Per (sector × regime) win-rate priors"),
        ("backtest/equity_backtest.py",           "Walk-forward replay engine"),
        ("reports/generate_cheat_sheet.py",       "This file ⟶ Strategy_Cheat_Sheet_Equity_Terminal.xlsx"),
        ("data/trades.json",                      "Live + closed positions (autosaved)"),
        ("data/cache/ohlcv_cache.db",             "SQLite OHLCV cache"),
        ("data/models/",                          "Per-symbol trained model binaries"),
        ("fyers_token.json",                      "Active Fyers OAuth token (gitignored)"),
        ("app.log",                               "Rotating app log (gitignored)"),
    ]
    write_table(ws, 3, ["File / Path", "Purpose"], rows,
                header_fill=ACCENT_BLUE,
                zebra=(LIGHT_BLUE, WHITE),
                col_widths=[42, 75])

sheet_files()


# ──────────────────────────────────────────────────────────────────
# 11. Charges Reference (NSE Equity)
# ──────────────────────────────────────────────────────────────────
def sheet_charges():
    ws = wb.create_sheet("11. Charges")
    style_title(ws, 1, 3, "NSE Equity Charges (built into P&L)", fill=ACCENT_YELLOW)
    style_subtitle(ws, 2, 3, "Used by trading/charge_calculator.py")

    rows = [
        ("STT — delivery",       f"{config.STT_DELIVERY_PCT}% (both sides)",
         "Securities Transaction Tax on delivery trades"),
        ("STT — intraday",       f"{config.STT_INTRADAY_PCT}% (sell side only)",
         "STT on intraday MIS trades"),
        ("Exchange charges",     f"{config.EXCHANGE_CHARGES_PCT}%",
         "NSE transaction charge"),
        ("GST",                  f"{config.GST_PCT}%",
         "On brokerage + exchange charges"),
        ("SEBI charges",         f"₹ {config.SEBI_CHARGES_PER_CRORE}/crore",
         "Statutory SEBI fee"),
        ("Stamp duty",           f"{config.STAMP_DUTY_PCT}% (buy side)",
         "Govt stamp duty on purchase"),
        ("Brokerage",            f"₹ {config.BROKERAGE_PER_ORDER}/order",
         "Flat per-order brokerage"),
        ("Order MPP buffer",     f"{config.ORDER_MPP_PCT}%",
         "Slippage protection on LIMIT orders (configurable from UI)"),
    ]
    write_table(ws, 4, ["Charge", "Rate", "Notes"], rows,
                header_fill=ACCENT_YELLOW,
                zebra=(LIGHT_YELLOW, WHITE),
                col_widths=[26, 30, 65])

sheet_charges()


# ──────────────────────────────────────────────────────────────────
# Save
# ──────────────────────────────────────────────────────────────────
out_path = os.path.join(os.path.dirname(__file__),
                        "Strategy_Cheat_Sheet_Equity_Terminal.xlsx")
wb.save(out_path)
print(f"[OK] Cheat sheet generated: {out_path}")
print(f"     Sheets: {len(wb.sheetnames)} -> {wb.sheetnames}")
