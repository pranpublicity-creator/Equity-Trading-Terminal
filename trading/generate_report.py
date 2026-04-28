"""
Report Generator — NSE Equity Trading Terminal
Generates XLSX trade analysis reports.

Sheet layout
============
  1. Trade History       — every closed position, 26 columns incl. MAE/MFE/Grade
  2. Summary             — headline stats + JournalEngine metrics (Sharpe, Sortino, CAGR…)
  3. Per Strategy        — win-rate, profit factor, MAE/MFE per strategy preset
  4. Per Pattern         — same breakdown by chart-pattern name
  5. Excursion Analysis  — per-symbol MAE/MFE distribution + suggested optimal SL
  6. Symbol Ranking      — composite rank score (40% Sharpe + 30% Win% + 30% PF)
"""
import logging
import os
from datetime import datetime

import config

logger = logging.getLogger(__name__)


# ── Datetime helpers ──────────────────────────────────────────────────────────

def _fmt_dt(ts):
    if not ts:
        return None
    try:
        return datetime.fromtimestamp(float(ts))
    except Exception:
        return None

def _date_str(ts):
    dt = _fmt_dt(ts)
    return dt.strftime("%d/%m/%Y") if dt else ""

def _time_str(ts):
    dt = _fmt_dt(ts)
    return dt.strftime("%H:%M:%S") if dt else ""


# ── Style constants (set after openpyxl is imported) ─────────────────────────

def _make_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    return dict(
        HDR_FONT   = Font(bold=True, color="FFFFFF", size=10),
        HDR_FILL   = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid"),
        SUB_FILL   = PatternFill(start_color="2E4E6E", end_color="2E4E6E", fill_type="solid"),
        PROF_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        LOSS_FILL  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        NEUT_FILL  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        MFE_FILL   = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        MAE_FILL   = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        THIN       = Border(
            left=Side(style="thin"),   right=Side(style="thin"),
            top=Side(style="thin"),    bottom=Side(style="thin"),
        ),
        BOLD       = Font(bold=True, size=10),
        BOLD_LG    = Font(bold=True, size=12),
        CENTER     = Alignment(horizontal="center", vertical="center"),
        LEFT       = Alignment(horizontal="left",   vertical="center"),
        WRAP       = Alignment(wrap_text=True, horizontal="center", vertical="center"),
    )


def _hdr(ws, row, col, value, s, width=None):
    """Write a styled header cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = s["HDR_FONT"]
    cell.fill      = s["HDR_FILL"]
    cell.alignment = s["CENTER"]
    cell.border    = s["THIN"]
    if width:
        ws.column_dimensions[cell.column_letter].width = width
    return cell


def _grade_fill(grade, s):
    """Return fill colour matching quality grade A–F."""
    from openpyxl.styles import PatternFill
    MAP = {
        "A": "375623", "B": "70AD47", "C": "FFEB9C",
        "D": "F4B942", "F": "FFC7CE",
    }
    clr = MAP.get((grade or "").upper(), "FFFFFF")
    return PatternFill(start_color=clr, end_color=clr, fill_type="solid")


# ── Main entry point ─────────────────────────────────────────────────────────

def generate_trade_report(trade_history, output_path=None):
    """Generate XLSX report from list of closed Position dicts.

    Args:
        trade_history: list of dict (from dataclasses.asdict(position))
        output_path:   optional file path; auto-generated if None

    Returns:
        output_path on success, None on failure.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        logger.error("openpyxl not installed — run: pip install openpyxl")
        return None

    if not trade_history:
        logger.warning("No trades to report")
        return None

    if output_path is None:
        os.makedirs(config.REPORTS_DIR, exist_ok=True)
        output_path = os.path.join(
            config.REPORTS_DIR,
            f"equity_trade_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        )

    s = _make_styles()
    wb = Workbook()

    closed = [t for t in trade_history if t.get("status", "CLOSED") == "CLOSED"]

    # ── Run JournalEngine once for advanced metrics ───────────────────────────
    _journal_report = {}
    try:
        # Build minimal Position-like objects the JournalEngine accepts
        from trading.journal_engine import JournalEngine

        class _FakePos:
            pass

        fake_positions = []
        for t in closed:
            p = _FakePos()
            for k, v in t.items():
                setattr(p, k, v)
            fake_positions.append(p)

        je = JournalEngine(total_capital=getattr(config, "TOTAL_TRADING_CAPITAL", 500000))
        _journal_report = je.compute(fake_positions)
    except Exception as e:
        logger.warning(f"JournalEngine unavailable for report enrichment: {e}")
        _journal_report = {}

    # ── Sheet 1: Trade History ────────────────────────────────────────────────
    _build_trade_history(wb.active, closed, s)

    # ── Sheet 2: Summary ─────────────────────────────────────────────────────
    _build_summary(wb.create_sheet("Summary"), closed, _journal_report, s)

    # ── Sheet 3: Per Strategy ─────────────────────────────────────────────────
    _build_per_group(wb.create_sheet("Per Strategy"), closed, s, group_key="strategy")

    # ── Sheet 4: Per Pattern ──────────────────────────────────────────────────
    _build_per_group(wb.create_sheet("Per Pattern"), closed, s, group_key="pattern_name",
                     fallback="ML_SIGNAL")

    # ── Sheet 5: Excursion Analysis ───────────────────────────────────────────
    _build_excursion(wb.create_sheet("Excursion Analysis"), closed, _journal_report, s)

    # ── Sheet 6: Symbol Ranking ───────────────────────────────────────────────
    _build_symbol_ranking(wb.create_sheet("Symbol Ranking"), _journal_report, s)

    wb.save(output_path)
    logger.info(f"Equity trade report saved: {output_path}")
    return output_path


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_trade_history(ws, closed, s):
    ws.title = "Trade History"

    # 26 columns
    headers = [
        ("Trade\nID",       7),
        ("Symbol",         14),
        ("TF",              6),
        ("Dir",             7),
        ("Pattern",        22),
        ("Conf\n%",         8),
        ("Grade",           7),
        ("Qual\nScore",     9),
        ("Entry\nDate",    12),
        ("Entry\nTime",    10),
        ("Exit\nDate",     12),
        ("Exit\nTime",     10),
        ("Entry\nPrice",   12),
        ("Exit\nPrice",    12),
        ("Qty",             7),
        ("Gross P&L\n(₹)", 14),
        ("Charges\n(₹)",   12),
        ("Net P&L\n(₹)",   13),
        ("R:R",             7),
        ("MFE\n%",          8),
        ("MAE\n%",          8),
        ("Bars",            7),
        ("Exit\nReason",   22),
        ("Regime",         18),
        ("Strategy",       18),
        ("Signal\nStr",    11),
    ]

    ws.row_dimensions[1].height = 32
    for c, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = s["HDR_FONT"]
        cell.fill      = s["HDR_FILL"]
        cell.alignment = s["WRAP"]
        cell.border    = s["THIN"]
        ws.column_dimensions[cell.column_letter].width = w

    for row, t in enumerate(closed, 2):
        net_pnl   = t.get("realized_pnl", 0) or 0
        charges   = t.get("charges", 0) or 0
        gross_pnl = net_pnl + charges
        grade     = t.get("quality_grade", "") or ""
        mfe       = t.get("mfe_pct", 0) or 0
        mae       = t.get("mae_pct", 0) or 0   # stored negative
        tf_raw    = str(t.get("timeframe", "15"))
        tf_lbl    = "5m" if tf_raw == "5" else "15m"

        values = [
            row - 1,
            t.get("symbol", "").replace("NSE:", "").replace("-EQ", ""),
            tf_lbl,
            t.get("direction", ""),
            (t.get("pattern_name", "") or "ML").replace("_", " "),
            round(t.get("signal_confidence", 0) or 0, 1),
            grade,
            round((t.get("quality_score", 0) or 0) * 100, 1),   # 0-1 → 0-100
            _date_str(t.get("entry_time")),
            _time_str(t.get("entry_time")),
            _date_str(t.get("exit_time")),
            _time_str(t.get("exit_time")),
            round(t.get("entry_price", 0) or 0, 2),
            round(t.get("exit_price",  0) or 0, 2),
            t.get("quantity", 0) or 0,
            round(gross_pnl, 2),
            round(charges,   2),
            round(net_pnl,   2),
            round(t.get("risk_reward", 0) or 0, 2),
            round(mfe, 3),
            round(abs(mae), 3),          # show absolute value (was negative)
            t.get("bars_in_trade", 0) or 0,
            t.get("exit_reason", ""),
            t.get("regime", ""),
            t.get("strategy", ""),
            t.get("signal_strength", ""),
        ]

        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border    = s["THIN"]
            cell.alignment = s["CENTER"]

            # Net P&L colour
            if c == 18:
                cell.number_format = "#,##0.00"
                cell.fill = s["PROF_FILL"] if net_pnl > 0 else s["LOSS_FILL"]
            # Gross P&L / Charges
            elif c in (16, 17):
                cell.number_format = "#,##0.00"
            # Grade cell — colour by letter
            elif c == 7 and grade:
                cell.fill = _grade_fill(grade, s)
                cell.font = s["BOLD"]
            # MFE column — light blue tint
            elif c == 20:
                cell.fill = s["MFE_FILL"]
                cell.number_format = "0.000"
            # MAE column (absolute) — light orange tint
            elif c == 21:
                cell.fill = s["MAE_FILL"]
                cell.number_format = "0.000"
            # Direction colour
            elif c == 4:
                from openpyxl.styles import Font as _F
                cell.font = _F(
                    bold=True, size=10,
                    color="375623" if v == "BUY" else "9C0006"
                )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _build_summary(ws, closed, journal, s):
    winners   = [t for t in closed if (t.get("realized_pnl") or 0) > 0]
    losers    = [t for t in closed if (t.get("realized_pnl") or 0) <= 0]
    total_pnl = sum(t.get("realized_pnl", 0) or 0 for t in closed)
    win_rate  = round(len(winners) / len(closed) * 100, 1) if closed else 0
    avg_win   = (sum(t.get("realized_pnl", 0) for t in winners) / len(winners)) if winners else 0
    avg_loss  = (abs(sum(t.get("realized_pnl", 0) for t in losers)) / len(losers)) if losers else 0
    wl_ratio  = round(avg_win / avg_loss, 2) if avg_loss > 0 else 0
    gross_wins = sum((t.get("realized_pnl", 0) or 0) for t in winners)
    gross_loss = abs(sum((t.get("realized_pnl", 0) or 0) for t in losers))
    pf         = round(gross_wins / gross_loss, 2) if gross_loss > 0 else 0

    # Excursion stats
    mfe_vals = [t.get("mfe_pct", 0) or 0 for t in closed]
    mae_vals = [abs(t.get("mae_pct", 0) or 0) for t in closed]
    win_mfe  = [t.get("mfe_pct", 0) or 0 for t in winners]
    avg_mfe  = round(sum(mfe_vals) / len(mfe_vals), 3) if mfe_vals else 0
    avg_mae  = round(sum(mae_vals) / len(mae_vals), 3) if mae_vals else 0

    # MFE capture: exit_pnl_pct / mfe_pct for each trade
    captures = []
    for t in closed:
        ep = t.get("entry_price", 0) or 0
        xp = t.get("exit_price",  0) or 0
        mfe = t.get("mfe_pct", 0) or 0
        if ep > 0 and mfe > 0:
            dir_pct = ((xp - ep) / ep * 100) if t.get("direction") == "BUY" else ((ep - xp) / ep * 100)
            captures.append(min(dir_pct / mfe * 100, 100))
    mfe_capture = round(sum(captures) / len(captures), 1) if captures else 0

    # Pull JournalEngine advanced metrics
    jsumm = journal.get("summary", {})

    # Build rows
    title_row = [("EQUITY TRADE REPORT — PERFORMANCE SUMMARY", None)]

    basic_rows = [
        ("Report Generated",    datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Total Capital",       f"₹ {getattr(config, 'TOTAL_TRADING_CAPITAL', 0):,.0f}"),
        ("", ""),
        ("── TRADE COUNTS ──", ""),
        ("Total Closed Trades",  len(closed)),
        ("Winners",              len(winners)),
        ("Losers",               len(losers)),
        ("Win Rate",             f"{win_rate}%"),
        ("", ""),
        ("── P&L ──", ""),
        ("Total Net P&L",        f"₹ {total_pnl:,.2f}"),
        ("Avg Net P&L / Trade",  f"₹ {total_pnl/len(closed):,.2f}" if closed else "₹ 0"),
        ("Avg Win",              f"₹ {avg_win:,.2f}"),
        ("Avg Loss",             f"₹ {avg_loss:,.2f}"),
        ("W/L Ratio",            f"{wl_ratio}×"),
        ("Profit Factor",        str(pf)),
        ("Gross Profit",         f"₹ {gross_wins:,.2f}"),
        ("Gross Loss",           f"₹ {gross_loss:,.2f}"),
        ("", ""),
        ("── RISK-ADJUSTED RETURNS ──", ""),
        ("Sharpe Ratio (ann.)",  jsumm.get("sharpe",  "--")),
        ("Sortino Ratio (ann.)", jsumm.get("sortino", "--")),
        ("CAGR",                 f"{jsumm.get('cagr_pct', 0):.2f}%" if jsumm.get("cagr_pct") is not None else "--"),
        ("Max Drawdown (₹)",     f"{jsumm.get('max_dd', 0):,.2f}" if jsumm.get("max_dd") is not None else "--"),
        ("", ""),
        ("── EXCURSION METRICS ──", ""),
        ("Avg MFE %",            f"{avg_mfe:.3f}%"),
        ("Avg MAE % (abs)",      f"{avg_mae:.3f}%"),
        ("MFE Capture %",        f"{mfe_capture:.1f}%"),
        ("Avg MFE % (winners)",  f"{sum(win_mfe)/len(win_mfe):.3f}%" if win_mfe else "--"),
        ("", ""),
        ("── QUALITY ──", ""),
    ]

    # Grade distribution
    grade_counts: dict = {}
    for t in closed:
        g = t.get("quality_grade", "?") or "?"
        grade_counts[g] = grade_counts.get(g, 0) + 1
    for g in sorted(grade_counts):
        basic_rows.append((f"  Grade {g}", f"{grade_counts[g]} trades"))

    # Write
    ws["A1"] = "EQUITY TRADE REPORT — PERFORMANCE SUMMARY"
    ws["A1"].font = s["BOLD_LG"]
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = s["LEFT"]

    for r, (label, value) in enumerate(basic_rows, 2):
        cell_l = ws.cell(row=r, column=1, value=label)
        cell_v = ws.cell(row=r, column=2, value=value)
        if label.startswith("──"):
            cell_l.font = s["BOLD"]
            cell_l.fill = s["SUB_FILL"]
            from openpyxl.styles import Font as _F
            cell_l.font = _F(bold=True, color="FFFFFF", size=10)
            ws.merge_cells(f"A{r}:B{r}")
        elif label:
            cell_l.font = s["BOLD"]
        if str(value).startswith("₹") and total_pnl != 0 and "P&L" in label:
            try:
                v_num = float(str(value).replace("₹ ", "").replace(",", ""))
                cell_v.fill = s["PROF_FILL"] if v_num > 0 else s["LOSS_FILL"]
            except Exception:
                pass

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22


def _build_per_group(ws, closed, s, group_key, fallback="unknown"):
    """Generic per-group breakdown sheet (works for Strategy and Pattern)."""
    title = "Strategy" if group_key == "strategy" else "Pattern"

    headers = [
        title,
        "Trades", "Winners", "Losers", "Win Rate %",
        "Total P&L (₹)", "Avg P&L (₹)", "Profit Factor",
        "Avg MFE %", "Avg MAE %",
        "Avg R:R", "Avg Conf %",
    ]
    widths  = [24, 8, 8, 8, 11, 14, 13, 13, 11, 11, 9, 11]

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        _hdr(ws, 1, c, h, s, width=w)

    groups: dict = {}
    for t in closed:
        key = t.get(group_key, "") or fallback
        if not key:
            key = fallback
        key = key.replace("_", " ")
        if key not in groups:
            groups[key] = {
                "trades": 0, "wins": 0, "losses": 0,
                "pnl": 0.0, "mfe": [], "mae": [], "rr": [], "conf": [],
            }
        pnl = t.get("realized_pnl", 0) or 0
        g   = groups[key]
        g["trades"] += 1
        g["wins"]   += 1 if pnl > 0 else 0
        g["losses"] += 1 if pnl <= 0 else 0
        g["pnl"]    += pnl
        g["mfe"].append(t.get("mfe_pct", 0) or 0)
        g["mae"].append(abs(t.get("mae_pct", 0) or 0))
        g["rr"].append(t.get("risk_reward", 0) or 0)
        g["conf"].append(t.get("signal_confidence", 0) or 0)

    for row, (grp, g) in enumerate(
        sorted(groups.items(), key=lambda x: x[1]["pnl"], reverse=True), 2
    ):
        wr   = round(g["wins"] / g["trades"] * 100, 1) if g["trades"] else 0
        avg  = round(g["pnl"] / g["trades"], 2) if g["trades"] else 0
        _grp_pnls = [
            t.get("realized_pnl", 0) or 0 for t in closed
            if (t.get(group_key, "") or fallback).replace("_", " ") == grp
        ]
        gw   = sum(v for v in _grp_pnls if v > 0)
        gl   = abs(sum(v for v in _grp_pnls if v <= 0))
        pf   = round(gw / gl, 2) if gl > 0 else 0
        avmfe = round(sum(g["mfe"]) / len(g["mfe"]), 3) if g["mfe"] else 0
        avmae = round(sum(g["mae"]) / len(g["mae"]), 3) if g["mae"] else 0
        avrr  = round(sum(g["rr"])  / len(g["rr"]),  2) if g["rr"]  else 0
        avcf  = round(sum(g["conf"])/ len(g["conf"]),1) if g["conf"] else 0

        vals = [grp, g["trades"], g["wins"], g["losses"], wr,
                round(g["pnl"], 2), avg, pf, avmfe, avmae, avrr, avcf]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = s["THIN"]
            cell.alignment = s["CENTER"]
            if c == 6:                          # Total P&L
                cell.number_format = "#,##0.00"
                cell.fill = s["PROF_FILL"] if v > 0 else s["LOSS_FILL"]
            elif c == 7:                        # Avg P&L
                cell.number_format = "#,##0.00"
            elif c == 9:                        # Avg MFE
                cell.fill = s["MFE_FILL"]
            elif c == 10:                       # Avg MAE
                cell.fill = s["MAE_FILL"]


def _build_excursion(ws, closed, journal, s):
    """Sheet 5: Per-symbol MAE/MFE excursion analysis + suggested SL."""
    ws.column_dimensions["A"].width = 16

    # Title
    ws["A1"] = "EXCURSION ANALYSIS — Max Adverse / Max Favourable Excursion by Symbol"
    ws["A1"].font = s["BOLD_LG"]
    ws.merge_cells("A1:K1")

    headers = [
        "Symbol", "Trades", "Win Rate %",
        "Avg MFE %", "Avg MAE %",
        "Best MFE %", "Worst MAE %",
        "MAE P50 %", "MAE P80 %",
        "Suggested SL %",
        "MFE Capture %",
    ]
    widths = [16, 8, 11, 11, 11, 12, 12, 11, 11, 14, 14]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        _hdr(ws, 2, c, h, s, width=w)

    # Pull from JournalEngine symbol_stats if available
    sym_stats = journal.get("symbol_stats", {})

    # Fall back to in-row computation if journal is empty
    if not sym_stats:
        from collections import defaultdict
        sym_stats_raw = defaultdict(lambda: {
            "trades": 0, "wins": 0,
            "mfe": [], "mae": [], "mfe_cap": [],
        })
        for t in closed:
            sym = t.get("symbol", "").replace("NSE:", "").replace("-EQ", "")
            ep  = t.get("entry_price", 0) or 0
            xp  = t.get("exit_price",  0) or 0
            mfe = t.get("mfe_pct", 0) or 0
            mae = abs(t.get("mae_pct", 0) or 0)
            pnl = t.get("realized_pnl", 0) or 0
            g   = sym_stats_raw[sym]
            g["trades"] += 1
            g["wins"]   += 1 if pnl > 0 else 0
            g["mfe"].append(mfe)
            g["mae"].append(mae)
            if ep > 0 and mfe > 0:
                dir_pct = ((xp - ep)/ep*100) if t.get("direction") == "BUY" else ((ep - xp)/ep*100)
                g["mfe_cap"].append(min(dir_pct / mfe * 100, 100))

        def _pct(lst, p):
            if not lst:
                return 0
            lst = sorted(lst)
            idx = (len(lst) - 1) * p / 100.0
            lo, hi = int(idx), min(int(idx) + 1, len(lst) - 1)
            return round(lst[lo] + (lst[hi] - lst[lo]) * (idx - lo), 3)

        sym_stats = {}
        for sym, g in sym_stats_raw.items():
            wr = round(g["wins"] / g["trades"] * 100, 1) if g["trades"] else 0
            win_mfe  = [m for t, m in zip(closed, g["mfe"])
                        if t.get("symbol","").replace("NSE:","").replace("-EQ","") == sym
                        and (t.get("realized_pnl",0) or 0) > 0]
            sym_stats[sym] = {
                "trades":           g["trades"],
                "win_rate":         wr,
                "avg_mfe_pct":      round(sum(g["mfe"]) / len(g["mfe"]), 3) if g["mfe"] else 0,
                "avg_mae_pct":      round(sum(g["mae"]) / len(g["mae"]), 3) if g["mae"] else 0,
                "best_mfe_pct":     round(max(g["mfe"]), 3) if g["mfe"] else 0,
                "worst_mae_pct":    round(max(g["mae"]), 3) if g["mae"] else 0,
                "mae_p50":          _pct(g["mae"], 50),
                "mae_p80":          _pct(g["mae"], 80),
                "suggested_sl_pct": _pct([m for t, m in zip(closed, g["mae"])
                                          if t.get("symbol","").replace("NSE:","").replace("-EQ","") == sym
                                          and (t.get("realized_pnl",0) or 0) > 0], 80),
                "mfe_capture_pct":  round(sum(g["mfe_cap"]) / len(g["mfe_cap"]), 1) if g["mfe_cap"] else 0,
            }

    # Write rows sorted by trade count desc
    for row, (sym, st) in enumerate(
        sorted(sym_stats.items(),
               key=lambda x: x[1].get("trades", x[1].get("total_trades", 0)),
               reverse=True), 3
    ):
        wr  = st.get("win_rate",         0) or 0
        slp = st.get("suggested_sl_pct", 0) or 0
        cap = st.get("mfe_capture_pct",  0) or 0
        n_t = st.get("trades", st.get("total_trades", 0))

        vals = [
            sym,
            n_t,
            round(wr, 1),
            round(st.get("avg_mfe_pct",              0) or 0, 3),
            round(abs(st.get("avg_mae_pct",          0) or 0), 3),
            round(st.get("best_mfe_pct",             0) or 0, 3),
            round(st.get("worst_mae_pct",            0) or 0, 3),
            round(st.get("mae_p50",                  0) or 0, 3),
            round(st.get("mae_p80",                  0) or 0, 3),
            round(slp if slp is not None else 0,         3),
            round(cap,                                   1),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border    = s["THIN"]
            cell.alignment = s["CENTER"]
            if c == 3:   # Win Rate
                cell.fill = (s["PROF_FILL"] if wr >= 60 else
                             s["NEUT_FILL"] if wr >= 45 else
                             s["LOSS_FILL"])
            elif c == 4:   # Avg MFE
                cell.fill = s["MFE_FILL"]
            elif c == 5:   # Avg MAE
                cell.fill = s["MAE_FILL"]
            elif c == 10:  # Suggested SL
                cell.fill = s["NEUT_FILL"]
                cell.font = s["BOLD"]
                if isinstance(v, float):
                    cell.number_format = "0.000\"%\""
            elif c == 11:  # MFE Capture
                cell.fill = (s["PROF_FILL"] if cap >= 70 else
                             s["NEUT_FILL"] if cap >= 50 else
                             s["LOSS_FILL"])

    # Freeze header rows
    ws.freeze_panes = "A3"


def _build_symbol_ranking(ws, journal, s):
    """Sheet 6: Composite symbol rank score from JournalEngine."""
    ws["A1"] = "SYMBOL RANKING — Composite Score (40% Sharpe + 30% Win Rate + 30% Profit Factor)"
    ws["A1"].font = s["BOLD_LG"]
    ws.merge_cells("A1:L1")

    headers = [
        "Rank", "Symbol", "Trades",
        "Win Rate %", "Profit Factor", "Sharpe",
        "MAE P80 %", "Avg MFE %", "Avg MAE %",
        "SL Suggest %", "Net P&L (₹)", "Score /100",
    ]
    widths = [7, 16, 8, 11, 13, 10, 11, 11, 11, 13, 14, 12]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        _hdr(ws, 2, c, h, s, width=w)

    ranking = journal.get("symbol_ranking", [])
    if not ranking:
        ws.cell(row=3, column=1, value="No symbol data — run more trades to populate ranking.").font = s["BOLD"]
        return

    for row, entry in enumerate(ranking, 3):
        rank  = row - 2
        score = entry.get("rank_score", 0) or 0
        wr    = entry.get("win_rate",   0) or 0
        pnl   = entry.get("total_pnl",  0) or 0

        vals = [
            rank,
            entry.get("symbol", ""),
            entry.get("trades", entry.get("total_trades", 0)),
            round(wr, 1),
            round(entry.get("profit_factor",    0) or 0, 2),
            round(entry.get("sharpe",           0) or 0, 2),
            round(entry.get("mae_p80",          0) or 0, 3),
            round(entry.get("avg_mfe_pct",      0) or 0, 3),
            round(abs(entry.get("avg_mae_pct",  0) or 0), 3),
            round(entry.get("suggested_sl_pct", 0) or 0, 3),
            round(pnl, 2),
            round(score, 1),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border    = s["THIN"]
            cell.alignment = s["CENTER"]

            # Rank medal colours (top 3)
            if c == 1:
                if rank == 1:
                    from openpyxl.styles import PatternFill as PF
                    cell.fill = PF(start_color="FFD700", end_color="FFD700", fill_type="solid")
                    cell.font = s["BOLD"]
                elif rank == 2:
                    from openpyxl.styles import PatternFill as PF
                    cell.fill = PF(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
                    cell.font = s["BOLD"]
                elif rank == 3:
                    from openpyxl.styles import PatternFill as PF
                    cell.fill = PF(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
                    cell.font = s["BOLD"]

            # Win Rate colour
            elif c == 4:
                cell.fill = (s["PROF_FILL"] if wr >= 60 else
                             s["NEUT_FILL"] if wr >= 45 else
                             s["LOSS_FILL"])
            # PF colour
            elif c == 5:
                pf_v = v or 0
                cell.fill = (s["PROF_FILL"] if pf_v >= 1.5 else
                             s["NEUT_FILL"] if pf_v >= 1.0 else
                             s["LOSS_FILL"])
            # Net P&L colour
            elif c == 11:
                cell.number_format = "#,##0.00"
                cell.fill = s["PROF_FILL"] if pnl > 0 else s["LOSS_FILL"]
            # Score — colour gradient green/amber/red
            elif c == 12:
                cell.font = s["BOLD"]
                cell.fill = (s["PROF_FILL"] if score >= 70 else
                             s["NEUT_FILL"] if score >= 45 else
                             s["LOSS_FILL"])
            elif c == 8:
                cell.fill = s["MFE_FILL"]
            elif c == 9:
                cell.fill = s["MAE_FILL"]

    ws.freeze_panes = "A3"
